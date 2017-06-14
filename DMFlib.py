import openpyxl,json,re,urllib,xlrd
from yahoo_finance import Share

def IncomeStatement(company):
	urlbase = 'http://financials.morningstar.com/ajax/ReportProcess4CSV.html?t='  # base url to talk with morningstar
	ISparameters = {'reportType': 'is', 'period': '12', 'dataType': 'A', 'order': 'desc', 'columnYear': '5',
					'number': '3'}  # parameters to build URL for Income Statement
	tagsIS = [('revenue',
			   'Revenue.*?,(\S*?),')]  # tags to be extracted from income statement with is matching regex (list of tuples)
	urlIS = urlbase + company + '&&' + urllib.urlencode(ISparameters)  # generate income statement url
	dataIS = urllib.urlopen(urlIS).read()  # get all necessary data
	fileIS=open('rawdataIS.txt','w')
	fileIS.write(dataIS)
	fileIS.close()
	companyinfo={}
	for tag,regex in tagsIS:
		end=regex.find('.')
		if regex[:end] in dataIS:
			companyinfo[tag]=float(re.findall(regex,dataIS)[0])
		else:
			companyinfo[tag]=0
	companyinfo['currency'] = re.findall('.*\. (\S{3}) in millions except per share data', dataIS)[
		0]  # get what currency the data is in
	return companyinfo

def BalanceSheet(company):
	urlbase = 'http://financials.morningstar.com/ajax/ReportProcess4CSV.html?t='  # base url to talk with morningstar
	BSparameters = {'reportType': 'bs', 'period': '12', 'dataType': 'A', 'order': 'desc', 'columnYear': '5',
					'number': '3'}  # parameters to build URL for Balance Sheet
	tagsBS = [('TCA', 'Total current assets.*?,(\S*?),'), ('TNCA', 'Total non-current assets.*?,(\S*?),'),
			  ('TCL', 'Total current liabilities.*?,(\S*?),'), ('TNCL', 'Total non-current liabilities.*?,(\S*?),'),
			  ('GW', 'Goodwill.*?,(\S*?),'), ('IA', 'Intangible assets.*?,(\S*?),'),
			  ('LTD', 'Long-term debt.*?,(\S*?),'),
			  ('OLTL', 'Other long-term liabilities.*?,(\S*?),')]  # tags to be extracted from balance sheet
	urlBS = urlbase + company + '&&' + urllib.urlencode(BSparameters)  # generate balance sheet url
	dataBS = urllib.urlopen(urlBS).read()
	fileBS = open('rawdataBS.txt', 'w')
	fileBS.write(dataBS)
	fileBS.close()
	companyinfo={}
	for tag,regex in tagsBS:
		end=regex.find('.')
		if regex[:end] in dataBS:
			try: companyinfo[tag]=float(re.findall(regex,dataBS)[0])
			except:
				companyinfo[tag]=0
				print 'warning: ', tag, ' was not a number.'
		else:
			companyinfo[tag]=0
	return companyinfo

def KeyRatios(company):
	urlKR = 'http://financials.morningstar.com/ajax/exportKR2CSV.html?t=' + company + '&'  # generate key ratios url
	dataKR = urllib.urlopen(urlKR).read()
	fileKR = open('rawdataKR.txt', 'w')
	fileKR.write(dataKR)
	fileKR.close()
	companyinfo = {}
	epsall = re.findall(
		'Earnings Per Share.*?,(\S*?),(\S*?),(\S*?),(\S*?),(\S*?),(\S*?),(\S*?),(\S*?),(\S*?),(\S*?),(\S*)',
		dataKR)  # get 10 years of eps
	i, addES, ES = 1, True, 0
	while i < 11:  # match eps1-10 with corresponding values and insert into companyinfo
		try:
			companyinfo['eps' + str(i)] = float(epsall[0][11 - i])
		except:
			companyinfo['eps' + str(i)] = 0
		if addES and companyinfo['eps' + str(i)] > 0:
			ES += 1
		else:
			addES = False
		i += 1
	companyinfo['ES'] = ES
	dividends = re.findall('Dividends.*?,(\S*?),(\S*?),(\S*?),(\S*?),(\S*?),(\S*?),(\S*?),(\S*?),(\S*?),(\S*?),(\S*?)',
						   dataKR)  # get all the dividends paid
	DR, i = 0, 9
	while i > 0:  # calculate dividend record
		try:
			if float(dividends[0][i]) > 0:
				DR += 1
				i -= 1
			else:
				break
		except:
			break
	companyinfo['DR'] = DR
	return companyinfo

def ExchangeRate(currency):
	if currency=='USD':EXR=1
	else:
		APIID = 'fe39ef6b876a482e874dba7bb29af7de'  # ID for url for exchange rate
		urlEXR = 'https://openexchangerates.org/api/latest.json?app_id=' + APIID  # build exchangerate DB URL
		dataEXR = urllib.urlopen(urlEXR).read()
		jsonEXR = json.loads(dataEXR)
		rates = jsonEXR.get('rates', None)
		EXR = rates.get(currency, None)  # get the exchange rate
	return EXR

def WriteToExcel(company,companyinfo,month,year):
	number, cell = 0, 0
	try:
		wb = openpyxl.load_workbook(month + ' ' + year + '.xlsx')
		ws = wb.get_sheet_by_name('sheet1')
	except:
		cell, number = None, 1
		print 'no existing file'
		wb = openpyxl.load_workbook('blankDMF.xlsx')
		ws = wb.get_sheet_by_name('sheet1')
	while cell is not None:
		number += 1
		cell = ws['C' + str(number * 3 + 1)].value
	ws['B' + str(number * 3 + 1)] = companyinfo['name']
	ws['C' + str(number * 3 + 1)] = company
	ws['F' + str(number * 3 + 1)] = companyinfo['EXR']
	ws['G' + str(number * 3 + 1)] = companyinfo['quote']
	ws['H' + str(number * 3 + 1)] = companyinfo['MC']
	ws['I' + str(number * 3 + 1)] = companyinfo['eps1']
	ws['I' + str(number * 3 + 2)] = companyinfo['eps2']
	ws['I' + str(number * 3 + 3)] = companyinfo['eps3']
	ws['J' + str(number * 3 + 1)] = companyinfo['revenue']
	ws['L' + str(number * 3 + 1)] = companyinfo['TCA']
	ws['L' + str(number * 3 + 2)] = companyinfo['TNCA']
	ws['N' + str(number * 3 + 1)] = companyinfo['TCL']
	ws['N' + str(number * 3 + 2)] = companyinfo['TNCL']
	ws['P' + str(number * 3 + 1)] = companyinfo['GW']
	ws['P' + str(number * 3 + 2)] = companyinfo['IA']
	ws['R' + str(number * 3 + 1)] = companyinfo['LTD']
	ws['R' + str(number * 3 + 2)] = companyinfo['OLTL']
	ws['U' + str(number * 3 + 1)] = companyinfo['eps8']
	ws['U' + str(number * 3 + 2)] = companyinfo['eps9']
	ws['U' + str(number * 3 + 3)] = companyinfo['eps10']
	ws['S' + str(number * 3 + 1)] = companyinfo['ES']
	ws['V' + str(number * 3 + 1)] = companyinfo['DR']
	wb.save(month + ' ' + year + '.xlsx')
	number += 1
	print 'added',companyinfo['name'],'to',month,year+'.xlsx'
	return;


def Retrieve52WeekLows(day,month):
    url = 'http://247wallst.com/investing/'
    webdata = urllib.urlopen(url).read()
    links = re.findall('href="(http:.*?-52-week-low-club/)"', webdata)
    date = re.findall('\d{4}/\d{2}/\d{2}', links[0])[0]
    if date[5:]!=month+'/'+day:
        print 'Error: There is no 52 week low article for',month+'/'+day,52'...yet'
        return;
    url = links[0]
    webdata = urllib.urlopen(url).read()
    tickersnyse = re.findall('\(NYSE: (\w+)\)', webdata)
    tickersnas = re.findall('\(NASDAQ: (\w+)\)', webdata)
    tickers = tickersnyse + tickersnas
    print '52 week low companies for', date, 'are', tickers
    return tickers



def GetListFromFile(name):
    list = []
    file1 = open(name + '.csv')
    for line in file1:
        try:company = re.findall('\w+', line)[0]
        except:continue
        list.append(company)
    if len(list) == 0: print 'Error: List is empty.'
    return list
>>>>>>> d29cacd63c1e126c7966616d2c5ccf3ad4a41529
	
def ReadFromExcel(file):
	ticker_file = xlrd.open_workbook(file)
	sheet1 = ticker_file.sheet_by_index(0)
	tickers = sheet1.col_values(0)
	return tickers

def GetInfo(company):
	companyinfo = {}
	dictIS = IncomeStatement(company)
	dictBS = BalanceSheet(company)
	dictKR = KeyRatios(company)
	dictionaries = [dictIS, dictBS, dictKR]
	companyinfo['EXR'] = ExchangeRate(dictIS['currency'])

	# price and market capitalization
	share = Share(company)
	companyinfo['quote'] = share.get_price()
	companyinfo['MC'] = float(share.get_market_cap()[:-1])
	companyinfo['name'] = share.get_name()

	for dictionary in dictionaries:  # merge dictionaries
		companyinfo.update(dictionary)

<<<<<<< HEAD
	return companyinfo
=======
    return companyinfo

>>>>>>> d29cacd63c1e126c7966616d2c5ccf3ad4a41529
